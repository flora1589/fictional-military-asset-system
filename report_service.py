import os
import csv
import io
from datetime import datetime, timezone
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models.mission import Mission
from app.models.asset import Asset
from app.models.maintenance import MaintenanceRecord
from app.models.inventory import InventoryItem
from app.models.request import AssetRequest
from app.models.user import User

REPORTS_DIR = "./generated_reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

def generate_pdf_report(report_type: str, data: List[Dict[str, Any]], title: str) -> str:
    filename = f"{report_type.lower()}_report_{int(datetime.now(timezone.utc).timestamp())}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=12
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=18
    )
    
    elements.append(Paragraph(title.upper(), title_style))
    elements.append(Paragraph(f"DEFENSE LOGISTICS CAPSTONE SIMULATOR • Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} UTC", subtitle_style))
    
    if data and len(data) > 0:
        headers = list(data[0].keys())
        table_data = [[h.replace("_", " ").title() for h in headers]]
        
        for row in data:
            row_values = []
            for h in headers:
                val = str(row.get(h, ''))
                # Truncate long strings for PDF rendering
                if len(val) > 40:
                    val = val[:37] + "..."
                row_values.append(val)
            table_data.append(row_values)
            
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No record data available for this report criteria.", styles['Normal']))
        
    doc.build(elements)
    return filepath

def generate_excel_report(report_type: str, data: List[Dict[str, Any]], title: str) -> str:
    filename = f"{report_type.lower()}_report_{int(datetime.now(timezone.utc).timestamp())}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    
    # Title Header
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{title.upper()} - DEFENSE LOGISTICS SYSTEM"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    if data and len(data) > 0:
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header.replace("_", " ").title()
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        for row_num, row_data in enumerate(data, 4):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(row_data.get(header, ''))
                
    wb.save(filepath)
    return filepath

def generate_csv_report(report_type: str, data: List[Dict[str, Any]]) -> str:
    filename = f"{report_type.lower()}_report_{int(datetime.now(timezone.utc).timestamp())}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    if not data:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            f.write("No data\n")
        return filepath
        
    headers = list(data[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        
    return filepath
